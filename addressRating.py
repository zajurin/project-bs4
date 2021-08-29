import openpyxl
import requests
import bs4
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from bs4 import BeautifulSoup
import re

import phonenumbers
from phonenumbers import carrier, timezone, geocoder

#******* AVOID BEING BLOCKED *******
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


#******* WORKING WITH EXCEL *******
chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myAddresses.xlsx"
choose_SHEET_Of_Your_Excel_File = 'Hoja1'

wb = load_workbook(chooseExcel_File, data_only=True)
sh = wb[choose_SHEET_Of_Your_Excel_File]

rowM = sh.max_row


for r in range(2, rowM+1):
    try:
        #READER OF EXCEL FILE
        letter_Of_Column_With_Name_Of_Companies = "C"
        number_Of_Row_With_Name_Of_Companies = r
        string_Of_number_Of_Row_With_Name_Of_Companies = str(r)
        concatenater_Column_With_Number = letter_Of_Column_With_Name_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies
        companyName = sh[concatenater_Column_With_Number].value 

        print(companyName)

        #CREATING URL FOR GOOGLE SEARCHING
        text = ("address of {0} headquarters ".format(companyName))
        myCurrentURL = 'https://google.com/search?q=' + text

        #ACCESING TO URL's INFO
        response = requests.get(myCurrentURL, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        #Look for this tag
        tagSearched = soup.find_all('div', attrs={'class': 'cXedhc'})
        if tagSearched:
            print('tag exists')
            for tag in tagSearched:
                companyAddress = tag.text.strip()
                copyOfCompanyAddress = companyAddress.read()
                print(copyOfCompanyAddress)
        else:
            print("Does not")        

            # #Where to put that info
            # letter_Of_Column_To_Put_Addres = "D"
            # concatenater_Row_With_Column = letter_Of_Column_To_Put_Addres + string_Of_number_Of_Row_With_Name_Of_Companies
            # sh[concatenater_Row_With_Column] = companyAddress
            # wb.save(filename = chooseExcel_File)

            # for match in phonenumbers.PhoneNumberMatcher(copyOfCompanyAddress, "GB"):
            #     print(match)
            

  
    except:
        print("Something went wrong")