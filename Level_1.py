import requests
import bs4
import openpyxl
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from bs4 import BeautifulSoup

import re

import urllib.request
import phonenumbers
import pycountry

from listOfCountries_and_codes import GetCountryCode

from Level_3_OOP import Level_3

#******* AVOID BEING BLOCKED *******
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


#******* WORKING WITH EXCEL *******

#where is located your excel file? 
#Dont forget to include its name file at the end of the path with its extension. 
global chooseExcel_File
chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myPhonenumbers.xlsx"

#What's the name of the sheet where you are going to work?
choose_SHEET_Of_Your_Excel_File = 'Hoja1'

# Read an existing workbook
wb = load_workbook(chooseExcel_File, data_only=True)
sh = wb[choose_SHEET_Of_Your_Excel_File]

rowM = sh.max_row

for r in range(2, rowM+1):# Range starts with 2 because the row position in Excel's file
    try:
        #READER OF EXCEL FILE

        #Name of Company
        global letter_Of_Column_With_Name_Of_Companies
        letter_Of_Column_With_Name_Of_Companies = "C"

        #Address of Company
        global letter_Of_Column_With_ADDRESS_Of_Companies
        letter_Of_Column_With_ADDRESS_Of_Companies = "D" 

        #Country of Company
        global letter_Of_Column_With_COUNTRY_Of_Companies
        letter_Of_Column_With_COUNTRY_Of_Companies = "E" 

        #Number of your ROW in Excel
        number_Of_Row_With_Name_Of_Companies = r

        # Must to be GLOBAL because it will be used in Reader_Constructor and will be inherited to another class instancies
        global string_Of_number_Of_Row_With_Name_Of_Companies         
        string_Of_number_Of_Row_With_Name_Of_Companies = str(r)#Converting "r" to string (which is type number) so that can be concatenated with variable "letter_Of_Column_With_Name_Of_Companies"

        #*************** CONCATENATING COLUMNS WITH ROWS ****************
        
        # NAME of Company IN EXCEL
        global concatenater_Column_With_Number_For_NAME
        concatenater_Column_With_Number_For_NAME = letter_Of_Column_With_Name_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies        

        # ADDRESS of Company IN EXCEL
        global concatenater_Column_With_Number_for_ADDRESS
        concatenater_Column_With_Number_for_ADDRESS = letter_Of_Column_With_ADDRESS_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies

        # COUNTRY of Company IN EXCEL
        global concatenater_Column_With_Number_for_COUNTRY
        concatenater_Column_With_Number_For_COUNTRY = letter_Of_Column_With_COUNTRY_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies

        # ****************** COMPLETED VALUES TO BE EXPORTED ****************** 
        global companyName
        companyName = sh[concatenater_Column_With_Number_For_NAME].value 
        
        global companyAddress
        companyAddress = sh[concatenater_Column_With_Number_for_ADDRESS].value
        if companyAddress is None:
            companyAddress = " headquarters"

        # Dynamic Country name from excel file
        global whichCountry
        whichCountry = sh[concatenater_Column_With_Number_For_COUNTRY].value


        # ************* Added for country codes *************
        # listOfCountries = pycountry.countries

        # nameOfCountry = listOfCountries.get(name=whichCountry)
        # your_Country_Code = nameOfCountry.alpha_2
        # # print("{0} \n".format(nameOfCountry))
        # # print('your code is ' + your_Country_Code)

        # # List of Contries
        # turning_countries_to_list = list(listOfCountries)

        # number = 0
        # for x in turning_countries_to_list:
        #     print("{0}".format(turning_countries_to_list[number].name))
        #     number += 1        

        print(companyName)#Tell me which Company was reviewed
        print(companyAddress)#Tell me the address of that one
        GetCountryCode(whichCountry)#Tell me the country in the cell

        class Reader_Constructor:
            def __init__(self, x_tagSearched):
                self.x_tagSearched = x_tagSearched

                for tag in self.x_tagSearched:
                    companyPhoneNumber = tag.text.strip()# here must to be only_Number instead of "tag.text.strip()"
                    letter_Of_Column_To_Put_Phone_Number = "F"
                    concatenater_Row_With_Column_For_Number = letter_Of_Column_To_Put_Phone_Number + string_Of_number_Of_Row_With_Name_Of_Companies
                    sh[concatenater_Row_With_Column_For_Number] = companyPhoneNumber
                    wb.save(filename = chooseExcel_File)
                    

        #CREATING URL FOR GOOGLE SEARCHING
        text = ("phone number of {0} {1} ".format(companyName, companyAddress))
        
        global myCurrentURL
        myCurrentURL = 'https://google.com/search?q=' + text

        #ACCESING TO URL's INFO
        response = requests.get(myCurrentURL, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        #Look for this tag
        tagSearched_Level_1 = soup.find_all('span', attrs={'class': 'mw31Ze'})# Level 1
        tagSearched_Level_2= soup.find_all('div', attrs={'class': 'liYKde g VjDLd'})# Level 2
        tagSearched_Level_3 = soup.find_all('h3', text = re.compile(companyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 3
        # tagSearched_3 = soup.findAll('h3', text = re.compile(companyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 3
        if tagSearched_Level_1:
            print("Level 1")
            Reader_Constructor(tagSearched_Level_1)

        elif tagSearched_Level_2:
            print("Level 2")
        elif tagSearched_Level_3:
            print("Level 3") 
            Level_3(headers, 'US', companyName, companyAddress, 'mw31Ze', 'LC20lb DKV0Md', 'tF2Cxc', 'span', chooseExcel_File, choose_SHEET_Of_Your_Excel_File, wb, sh, rowM, string_Of_number_Of_Row_With_Name_Of_Companies)
        else:
            print('ELSE')

    except Exception as e:
        print(e)
        print("Something went wrong")
