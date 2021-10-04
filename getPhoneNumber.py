import openpyxl
import requests
import bs4
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from bs4 import BeautifulSoup

import re

import urllib.request
import phonenumbers


#******* AVOID BEING BLOCKED *******
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


#******* WORKING WITH EXCEL *******
chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myPhonenumbers.xlsx"
choose_SHEET_Of_Your_Excel_File = 'Hoja1'

wb = load_workbook(chooseExcel_File, data_only=True)
sh = wb[choose_SHEET_Of_Your_Excel_File]

rowM = sh.max_row

whichCountry = "US"

for r in range(2, rowM+1):
    try:
        #READER OF EXCEL FILE

        #Company Name
        letter_Of_Column_With_Name_Of_Companies = "C"
        letter_Of_Column_With_ADDRESS_Of_Companies = "D"        
        number_Of_Row_With_Name_Of_Companies = r

        # Must to be GLOBAL because it will be used in Reader_Constructor and will be inherited to another class instancies
        global string_Of_number_Of_Row_With_Name_Of_Companies 
        
        string_Of_number_Of_Row_With_Name_Of_Companies = str(r)

        concatenater_Column_With_Number_For_NAME = letter_Of_Column_With_Name_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies
        concatenater_Column_With_Number_for_ADDRESS = letter_Of_Column_With_ADDRESS_Of_Companies + string_Of_number_Of_Row_With_Name_Of_Companies
        global companyName
        companyName = sh[concatenater_Column_With_Number_For_NAME].value 
        
        global companyAddress
        companyAddress = sh[concatenater_Column_With_Number_for_ADDRESS].value
        if companyAddress is None:
            companyAddress = " headquarters"

        print(companyName)#Tell me which Company was reviewed
        print(companyAddress)#Tell me the address of that one

        class Reader_Constructor:
            def __init__(self, x_tagSearched):
                self.x_tagSearched = x_tagSearched

                for tag in self.x_tagSearched:
                    companyPhoneNumber = tag.text.strip()
                    letter_Of_Column_To_Put_Phone_Number = "E"
                    concatenater_Row_With_Column_For_Number = letter_Of_Column_To_Put_Phone_Number + string_Of_number_Of_Row_With_Name_Of_Companies
                    sh[concatenater_Row_With_Column_For_Number] = companyPhoneNumber
                    wb.save(filename = chooseExcel_File)

                

        #CREATING URL FOR GOOGLE SEARCHING
        text = ("phone number of {0} {1} ".format(companyName, companyAddress))
        global myCurrentURL
        myCurrentURL = 'https://google.com/search?q=' + text

        #ACCESING TO URL's INFO
        response = requests.get(myCurrentURL, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        #Look for this tag
        tagSearched_Level_1 = soup.find_all('span', attrs={'class': 'mw31Ze'})# Level 1
        tagSearched_Level_2= soup.find_all('div', attrs={'class': 'liYKde g VjDLd'})# Level 2
        tagSearched_Level_3 = soup.find_all('h3', text = re.compile(companyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 3
        # tagSearched_3 = soup.findAll('h3', text = re.compile(companyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 3
        if tagSearched_Level_1:
            print("Level 1")
            Reader_Constructor(tagSearched_Level_1)

            # for tag in tagSearched_Level_1:
            #     companyPhoneNumber = tag.text.strip()
            #     letter_Of_Column_To_Put_Phone_Number = "E"
            #     concatenater_Row_With_Column_For_Number = letter_Of_Column_To_Put_Phone_Number + string_Of_number_Of_Row_With_Name_Of_Companies
            #     sh[concatenater_Row_With_Column_For_Number] = companyPhoneNumber
            #     wb.save(filename = chooseExcel_File)
        elif tagSearched_Level_2:
            print("Level 2")
            execfile('Level_2.py')
        elif tagSearched_Level_3:
            print("Level 3")         
        else:
            print('ELSE')     
                           
  
    except:
        print("Something went wrong")

        
